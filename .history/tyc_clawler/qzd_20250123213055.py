from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
import time
import pandas as pd
from openpyxl import load_workbook

# 初始化 WebDriver
driver = webdriver.Edge()  # 使用 Edge 浏览器

# 输入企业名称
company_name = "搜索企业"

# 打开网页
url = f"https://www.qizhidao.com/check?searchKey={company_name}&tagNum=0&scPageTitle=PC%E7%AB%99%E4%B8%BB%E7%AB%99&searchModeType=8&fromRoutePage=check&businessSourceDzlQuery="
driver.get(url)

# 最大化窗口
driver.maximize_window()

# 等待页面加载
time.sleep(5)

# 点击搜索结果中的第一个企业链接
try:
    search_result = WebDriverWait(driver, 30).until(
        EC.visibility_of_element_located((By.XPATH, "//div[@id='searchItem']//a/span/em"))
    )
    search_result.click()
    time.sleep(2)
except Exception as e:
    print(f"点击搜索结果失败：{e}")

# 获取当前页面 URL
current_url = driver.current_url

# 提取企业信息
try:
    unified_social_credit_code = WebDriverWait(driver, 30).until(
        EC.visibility_of_element_located((By.XPATH, "//div[@id='businessInfo_num']//p[2]/div/span/span"))
    ).text
except Exception as e:
    unified_social_credit_code = "未找到统一社会信用代码"
    print(f"获取统一社会信用代码失败：{e}")

try:
    former_name = WebDriverWait(driver, 30).until(
        EC.visibility_of_element_located((By.XPATH, "//div[@id='businessInfo_num']//p[2]/span"))
    ).text
except Exception as e:
    former_name = "未找到曾用名"
    print(f"获取曾用名失败：{e}")

try:
    legal_representative = WebDriverWait(driver, 30).until(
        EC.visibility_of_element_located((By.XPATH, "//div[@id='businessInfo_num']//span[2]/a"))
    ).text
except Exception as e:
    legal_representative = "未找到法定代表人"
    print(f"获取法定代表人失败：{e}")

try:
    registered_capital = WebDriverWait(driver, 30).until(
        EC.visibility_of_element_located((By.XPATH, "//div[@id='businessInfo_num']//div[3]/div/div[2]/span/span"))
    ).text
except Exception as e:
    registered_capital = "未找到注册资本"
    print(f"获取注册资本失败：{e}")

try:
    organization_code = WebDriverWait(driver, 30).until(
        EC.visibility_of_element_located((By.XPATH, "//div[@id='businessInfo_num']//div[4]/div/div[2]/span/div/span/span"))
    ).text
except Exception as e:
    organization_code = "未找到组织机构代码"
    print(f"获取组织机构代码失败：{e}")

# 关闭浏览器
driver.quit()

# 将数据写入 Excel 文件
file_path = f"F:\\rpa\\企知道\\{company_name}_信息表.xlsx"
try:
    # 如果文件已存在，加载文件；否则创建新文件
    try:
        workbook = load_workbook(file_path)
    except FileNotFoundError:
        workbook = pd.ExcelWriter(file_path, engine='openpyxl').book

    sheet_name = "企业信息"
    if sheet_name not in workbook.sheetnames:
        workbook.create_sheet(sheet_name)
    sheet = workbook[sheet_name]

    # 写入数据
    sheet['B1'] = company_name
    sheet['B2'] = unified_social_credit_code
    sheet['B3'] = former_name
    sheet['B4'] = legal_representative
    sheet['B5'] = registered_capital
    sheet['B6'] = organization_code

    # 保存文件
    workbook.save(file_path)
    print(f"数据已成功写入文件：{file_path}")
except Exception as e:
    print(f"写入 Excel 文件失败：{e}")